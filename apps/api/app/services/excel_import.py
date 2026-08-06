"""
Excel-импорт: парсинг 7-вкладочного шаблона, валидация, создание в БД.
"""
import io
import uuid
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Optional

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.operation import Operation, OperationDependency, OperationResource
from app.models.project import Project
from app.models.resource import Resource
from app.models.production_order import ProductionOrder


# ── mapping enums: Russian → DB value ──────────────────────────────────
MODE_MAP = {
    "производство": "production", "строительство": "construction",
    "production": "production", "construction": "construction",
    "": "production",
}

BOOL_MAP = {
    "да": True, "true": True, "1": True, 1: True,
    "нет": False, "false": False, "0": False, 0: False,
    "": False, None: False,
}

NODE_TYPE_MAP = {
    "сборка": "assembly", "assembly": "assembly",
    "полуфабрикат": "semi_finished", "semi_finished": "semi_finished",
    "материал": "material", "material": "material",
    "фантом": "phantom", "phantom": "phantom",
}

RESOURCE_TYPE_MAP = {
    "оборудование": "equipment", "equipment": "equipment",
    "персонал": "labor", "labor": "labor",
    "бригада": "team", "team": "team",
    "инструмент": "tool", "tool": "tool",
    "транспорт": "vehicle", "vehicle": "vehicle",
}

PRIORITY_MAP = {
    "низкий": "low", "low": "low",
    "обычный": "normal", "normal": "normal",
    "высокий": "high", "high": "high",
    "критичный": "critical", "critical": "critical",
    "": "normal",
}


def _cell_str(row: tuple, idx: int) -> str:
    v = row[idx] if idx < len(row) else None
    return str(v).strip() if v is not None else ""


def _cell_int(row: tuple, idx: int) -> Optional[int]:
    s = _cell_str(row, idx)
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _cell_decimal(row: tuple, idx: int) -> Optional[Decimal]:
    s = _cell_str(row, idx)
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def _cell_date(row: tuple, idx: int) -> Optional[date]:
    v = row[idx] if idx < len(row) else None
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"]:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _find_sheet(wb, names: list[str], required: bool = False):
    for name in names:
        if name in wb.sheetnames:
            return wb[name]
    return None


class ImportResult:
    def __init__(self):
        self.created: dict = {}
        self.warnings: list[str] = []
        self.errors: list[str] = []

    @property
    def ok(self) -> bool:
        return len(self.errors) == 0


async def import_excel(
    file_bytes: bytes,
    tenant_id: uuid.UUID,
    user_id: uuid.UUID,
    db: AsyncSession,
) -> ImportResult:
    result = ImportResult()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # ── 1. Настройки ────────────────────────────────────────────────
    ws = _find_sheet(wb, ["1-Настройки", "Настройки"])
    if ws is None:
        result.errors.append("Вкладка «1-Настройки» не найдена")
        return result
    rows = list(ws.iter_rows(values_only=True))
    settings = {}
    for r in rows[1:]:
        k = _cell_str(r, 0)
        if k:
            settings[k.rstrip(" *")] = _cell_str(r, 1)

    project_name = settings.get("Название проекта", "").strip()
    if not project_name:
        result.errors.append("Название проекта не заполнено")
        return result

    mode = MODE_MAP.get(settings.get("Режим", "").strip().lower(), "production")
    uses_phases = BOOL_MAP.get(settings.get("Использовать этапы", "").strip().lower(), False)
    methods_raw = settings.get("Методы расчёта", "").strip()
    country = settings.get("Страна", "RU").strip() or "RU"

    # ── 2. Проект ───────────────────────────────────────────────────
    project = Project(
        tenant_id=tenant_id,
        name=project_name,
        status="draft",
        mode="quick",
        default_method="cpm",
        country_code=country,
        created_by=user_id,
    )
    db.add(project)
    await db.flush()
    result.created["project"] = str(project.id)

    # ── 3. Ресурсы ──────────────────────────────────────────────────
    res_map: dict[str, uuid.UUID] = {}
    ws_res = _find_sheet(wb, ["4-Ресурсы", "Ресурсы"])
    if ws_res:
        for row in list(ws_res.iter_rows(values_only=True))[1:]:
            name = _cell_str(row, 1)
            if not name:
                continue
            rtype = RESOURCE_TYPE_MAP.get(_cell_str(row, 2).lower(), "equipment")
            available = float(_cell_str(row, 4)) if _cell_str(row, 4) else 1.0
            unit = _cell_str(row, 5) or "pcs"
            res = Resource(
                tenant_id=tenant_id,
                name=name,
                resource_type=rtype,
                capacity_per_unit=Decimal(str(available)),
                unit=unit,
            )
            db.add(res)
            await db.flush()
            res_map[name] = res.id

    async def _get_or_create_resource(name: str) -> Optional[uuid.UUID]:
        if not name:
            return None
        if name in res_map:
            return res_map[name]
        res = Resource(
            tenant_id=tenant_id,
            name=name,
            resource_type="equipment",
            capacity_per_unit=Decimal("1"),
        )
        db.add(res)
        await db.flush()
        res_map[name] = res.id
        return res.id

    # ── 4. Заказы ───────────────────────────────────────────────────
    ws_orders = _find_sheet(wb, ["2-Заказы", "Заказы"])
    if ws_orders is None:
        result.errors.append("Вкладка «2-Заказы» не найдена")
        return result
    orders_created: list[tuple[str, uuid.UUID]] = []
    for row in list(ws_orders.iter_rows(values_only=True))[1:]:
        oid = _cell_str(row, 0)
        product = _cell_str(row, 1)
        qty = _cell_decimal(row, 3)
        if not oid or not product or qty is None:
            continue
        start_dt = _cell_date(row, 4)
        due_dt = _cell_date(row, 5)
        priority = PRIORITY_MAP.get(_cell_str(row, 6).lower(), "normal")
        client = _cell_str(row, 7)
        order = ProductionOrder(
            tenant_id=tenant_id,
            project_id=project.id,
            ext_id=oid,
            quantity=qty,
            start_date=start_dt,
            due_date=due_dt,
            priority=priority,
            client=client or None,
            status="draft",
        )
        db.add(order)
        await db.flush()
        orders_created.append((oid, order.id))

    if not orders_created:
        result.errors.append("Нет ни одного заказа")
        return result

    # ── 5. Маршруты ─────────────────────────────────────────────────
    ws_routes = _find_sheet(wb, ["5-Маршруты", "Маршруты"])
    if ws_routes is None:
        result.errors.append("Вкладка «5-Маршруты» не найдена")
        return result
    route_rows = list(ws_routes.iter_rows(values_only=True))
    if len(route_rows) < 2:
        result.errors.append("Вкладка «Маршруты» пуста")
        return result

    ops_data: list[dict] = []
    for row in route_rows[1:]:
        seq = _cell_int(row, 1)
        name = _cell_str(row, 2)
        dur = _cell_decimal(row, 6)
        if not seq or not name or dur is None:
            result.warnings.append(f"Пропущена строка: нет № оп. или названия или длительности")
            continue
        res_name = _cell_str(row, 3)
        pred_seq = _cell_int(row, 7)
        add_mat = _cell_str(row, 8)
        add_qty = _cell_decimal(row, 9)
        yrate = _cell_decimal(row, 10) or Decimal("1.0")

        res_id = await _get_or_create_resource(res_name) if res_name else None
        op = Operation(
            tenant_id=tenant_id,
            project_id=project.id,
            name=name,
            duration_base=dur,
            duration_unit="hour",
            yield_rate=yrate,
            operation_type="production",
        )
        db.add(op)
        await db.flush()
        # Связь с ресурсом
        if res_id:
            opr = OperationResource(
                operation_id=op.id,
                resource_id=res_id,
                role="primary",
            )
            db.add(opr)
        ops_data.append({"seq": seq, "id": op.id, "pred_seq": pred_seq})

    # Расставляем зависимости через OperationDependency
    sorted_ops = sorted(ops_data, key=lambda x: x["seq"])
    for i, opd in enumerate(sorted_ops):
        pred_seq = opd["pred_seq"]
        if pred_seq:
            for other in sorted_ops:
                if other["seq"] == pred_seq:
                    dep = OperationDependency(
                        predecessor_id=other["id"],
                        successor_id=opd["id"],
                        dependency_type="FS",
                    )
                    db.add(dep)
                    break
        elif i > 0:
            # Авто-последовательность
            prev = sorted_ops[i - 1]
            dep = OperationDependency(
                predecessor_id=prev["id"],
                successor_id=opd["id"],
                dependency_type="FS",
            )
            db.add(dep)

    await db.commit()

    result.created["orders"] = [oid for oid, _ in orders_created]
    result.created["operations"] = len(ops_data)
    result.created["resources"] = len(res_map)
    return result
