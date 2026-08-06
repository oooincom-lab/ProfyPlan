"""
Excel Export — выгрузка отчётов проекта в .xlsx.

Листы:
  1. Ганта — операции с ES/EF/LS/LF, цветовое кодирование критического пути
  2. CPM — расчётная таблица с формулами
  3. Ресурсы — загрузка ресурсов
  4. Сводка — PERT / Monte Carlo (если есть данные)
"""
import io
from dataclasses import dataclass
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# --- Styles ---
HEADER_FILL = None  # будет инициализирован
HEADER_FONT = None
CRITICAL_FILL = None
WARNING_FILL = None
NORMAL_FILL = None
THIN_BORDER = None
WRAP_ALIGN = None
CENTER_ALIGN = None


def _init_styles():
    global HEADER_FILL, HEADER_FONT, CRITICAL_FILL, WARNING_FILL, NORMAL_FILL
    global THIN_BORDER, WRAP_ALIGN, CENTER_ALIGN
    if HEADER_FILL is not None:
        return
    HEADER_FILL = PatternFill(start_color="1E1E2A", end_color="1E1E2A", fill_type="solid")
    HEADER_FONT = Font(name="Segoe UI", bold=True, color="E0E0E8", size=10)
    CRITICAL_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    NORMAL_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin", color="D0D0D8"),
        right=Side(style="thin", color="D0D0D8"),
        top=Side(style="thin", color="D0D0D8"),
        bottom=Side(style="thin", color="D0D0D8"),
    )
    WRAP_ALIGN = Alignment(wrap_text=True, vertical="center")
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def _style_header(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN


def _style_data(ws, start_row: int, end_row: int, cols: int):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
            cell.font = Font(name="Segoe UI", size=9)


def _auto_width(ws, cols: int, min_width: int = 8, max_width: int = 40):
    for c in range(1, cols + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c):
            for cell in row:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def build_export_excel(
    operations: list[dict],
    dependencies: list[dict],
    resources: Optional[list[dict]] = None,
    project_name: str = "Проект",
    include_pert: bool = False,
    pert_data: Optional[dict] = None,
    monte_carlo_data: Optional[dict] = None,
) -> bytes:
    """
    Генерирует Excel-файл с несколькими листами.

    operations: [{id, name, duration_base, early_start, early_finish, late_start, late_finish, slack, is_critical}, ...]
    dependencies: [{predecessor_id, successor_id, dependency_type, lag_time}, ...]
    resources: [{name, load_percent, bottleneck_level, assigned_operations}, ...]
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl не установлен")

    _init_styles()
    wb = Workbook()

    # --- Sheet 1: Ганта ---
    ws_gantt = wb.active
    ws_gantt.title = "Ганта"

    headers_gantt = [
        "№", "Операция", "Длит. (ч)", "ES", "EF", "LS", "LF",
        "Slack (ч)", "Крит.", "Предшественники", "Ресурсы"
    ]
    for c, h in enumerate(headers_gantt, 1):
        ws_gantt.cell(row=1, column=c, value=h)
    _style_header(ws_gantt, 1, len(headers_gantt))

    # Сортируем по ES
    ops = sorted(operations, key=lambda o: o.get("early_start", 0))

    # Словарь предшественников
    pred_map: dict[str, list] = {}
    for d in dependencies:
        pred_map.setdefault(d["successor_id"], []).append(d["predecessor_id"])

    # Словарь имён
    op_names: dict[str, str] = {o["id"]: o.get("name", o["id"][:8]) for o in operations}

    for i, op in enumerate(ops):
        r = i + 2
        ws_gantt.cell(row=r, column=1, value=i + 1)
        ws_gantt.cell(row=r, column=2, value=op.get("name", op["id"][:8]))
        ws_gantt.cell(row=r, column=3, value=round(op.get("duration_base", 0), 1))
        ws_gantt.cell(row=r, column=4, value=round(op.get("early_start", 0), 1))
        ws_gantt.cell(row=r, column=5, value=round(op.get("early_finish", 0), 1))
        ws_gantt.cell(row=r, column=6, value=round(op.get("late_start", 0), 1))
        ws_gantt.cell(row=r, column=7, value=round(op.get("late_finish", 0), 1))
        ws_gantt.cell(row=r, column=8, value=round(op.get("slack", 0), 1))
        ws_gantt.cell(row=r, column=9, value="✓" if op.get("is_critical") else "")
        preds = pred_map.get(op["id"], [])
        ws_gantt.cell(row=r, column=10, value=", ".join(op_names.get(p, p[:8]) for p in preds))

        # Подсветка критического пути
        if op.get("is_critical"):
            for c in range(1, len(headers_gantt) + 1):
                ws_gantt.cell(row=r, column=c).fill = CRITICAL_FILL

    _style_data(ws_gantt, 2, len(ops) + 1, len(headers_gantt))
    _auto_width(ws_gantt, len(headers_gantt))

    # --- Sheet 2: CPM ---
    ws_cpm = wb.create_sheet("CPM")

    headers_cpm = [
        "Операция", "Длит. (ч)", "ES", "EF", "LS", "LF",
        "Slack", "Крит.", "Ранний старт (формула)", "Ранний финиш (формула)"
    ]
    for c, h in enumerate(headers_cpm, 1):
        ws_cpm.cell(row=1, column=c, value=h)
    _style_header(ws_cpm, 1, len(headers_cpm))

    for i, op in enumerate(ops):
        r = i + 2
        ws_cpm.cell(row=r, column=1, value=op.get("name", op["id"][:8]))
        ws_cpm.cell(row=r, column=2, value=round(op.get("duration_base", 0), 1))
        ws_cpm.cell(row=r, column=3, value=round(op.get("early_start", 0), 1))
        ws_cpm.cell(row=r, column=4, value=round(op.get("early_finish", 0), 1))
        ws_cpm.cell(row=r, column=5, value=round(op.get("late_start", 0), 1))
        ws_cpm.cell(row=r, column=6, value=round(op.get("late_finish", 0), 1))
        ws_cpm.cell(row=r, column=7, value=round(op.get("slack", 0), 1))
        ws_cpm.cell(row=r, column=8, value="✓" if op.get("is_critical") else "")
        # Формулы
        ws_cpm.cell(row=r, column=9, value=round(op.get("early_start", 0), 1))
        ws_cpm.cell(row=r, column=10, value=round(op.get("early_finish", 0), 1))

        if op.get("is_critical"):
            for c in range(1, len(headers_cpm) + 1):
                ws_cpm.cell(row=r, column=c).fill = CRITICAL_FILL

    _style_data(ws_cpm, 2, len(ops) + 1, len(headers_cpm))

    # Итоги
    total_row = len(ops) + 2
    ws_cpm.cell(row=total_row, column=1, value="Итого (makespan)").font = Font(bold=True)
    ws_cpm.cell(row=total_row, column=5, value=round(max(
        (o.get("late_finish", 0) for o in ops), default=0
    ), 1)).font = Font(bold=True)

    _auto_width(ws_cpm, len(headers_cpm))

    # --- Sheet 3: Ресурсы ---
    if resources:
        ws_res = wb.create_sheet("Ресурсы")

        headers_res = [
            "Ресурс", "Загрузка %", "Уровень", "Операций", "Рекомендации"
        ]
        for c, h in enumerate(headers_res, 1):
            ws_res.cell(row=1, column=c, value=h)
        _style_header(ws_res, 1, len(headers_res))

        for i, res in enumerate(resources):
            r = i + 2
            ws_res.cell(row=r, column=1, value=res.get("name", "?"))
            load = res.get("load_percent", 0)
            ws_res.cell(row=r, column=2, value=round(load, 1))
            ws_res.cell(row=r, column=3, value=res.get("bottleneck_level", "normal"))
            ws_res.cell(row=r, column=4, value=res.get("assigned_operations", 0))
            recs = res.get("recommendations", [])
            ws_res.cell(row=r, column=5, value="; ".join(recs) if recs else "")

            # Цветовое кодирование
            if load > 95:
                for c in range(1, len(headers_res) + 1):
                    ws_res.cell(row=r, column=c).fill = CRITICAL_FILL
            elif load > 80:
                for c in range(1, len(headers_res) + 1):
                    ws_res.cell(row=r, column=c).fill = WARNING_FILL
            elif load > 0:
                for c in range(1, len(headers_res) + 1):
                    ws_res.cell(row=r, column=c).fill = NORMAL_FILL

        _style_data(ws_res, 2, len(resources) + 1, len(headers_res))
        _auto_width(ws_res, len(headers_res))

    # --- Sheet 4: PERT (опционально) ---
    if pert_data:
        ws_pert = wb.create_sheet("PERT")

        headers_pert = [
            "Параметр", "Значение"
        ]
        for c, h in enumerate(headers_pert, 1):
            ws_pert.cell(row=1, column=c, value=h)
        _style_header(ws_pert, 1, len(headers_pert))

        pert_rows = [
            ("Ожидаемая длительность (ч)", pert_data.get("total_expected", 0)),
            ("Стандартное отклонение (ч)", pert_data.get("total_std_dev", 0)),
            ("Дисперсия", pert_data.get("total_variance", 0)),
            ("68% дов. интервал (низ)", pert_data.get("confidence_68", {}).get("low", 0)),
            ("68% дов. интервал (верх)", pert_data.get("confidence_68", {}).get("high", 0)),
            ("95% дов. интервал (низ)", pert_data.get("confidence_95", {}).get("low", 0)),
            ("95% дов. интервал (верх)", pert_data.get("confidence_95", {}).get("high", 0)),
            ("Операций на критпути", len(pert_data.get("critical_path", []))),
        ]

        for i, (param, val) in enumerate(pert_rows):
            r = i + 2
            ws_pert.cell(row=r, column=1, value=param)
            ws_pert.cell(row=r, column=2, value=val if isinstance(val, (int, float)) else str(val))

        _style_data(ws_pert, 2, len(pert_rows) + 1, len(headers_pert))
        _auto_width(ws_pert, len(headers_pert))

    # --- Sheet 5: Monte Carlo (опционально) ---
    if monte_carlo_data:
        ws_mc = wb.create_sheet("Monte Carlo")

        headers_mc = [
            "Метрика", "Значение"
        ]
        for c, h in enumerate(headers_mc, 1):
            ws_mc.cell(row=1, column=c, value=h)
        _style_header(ws_mc, 1, len(headers_mc))

        mc_rows = [
            ("Итераций", monte_carlo_data.get("iterations", 0)),
            ("Детерминированная длит. (ч)", monte_carlo_data.get("deterministic_duration", 0)),
            ("Среднее (ч)", monte_carlo_data.get("mean", 0)),
            ("Стандартное отклонение (ч)", monte_carlo_data.get("std_dev", 0)),
            ("Минимум (ч)", monte_carlo_data.get("min_duration", 0)),
            ("Максимум (ч)", monte_carlo_data.get("max_duration", 0)),
            ("P50 (медиана)", monte_carlo_data.get("p50", 0)),
            ("P80", monte_carlo_data.get("p80", 0)),
            ("P90", monte_carlo_data.get("p90", 0)),
            ("P95", monte_carlo_data.get("p95", 0)),
            ("P99", monte_carlo_data.get("p99", 0)),
        ]
        for p in monte_carlo_data.get("percentiles", {}):
            mc_rows.append((f"P{p.get('level', '?')}", p.get("value", 0)))

        for i, (param, val) in enumerate(mc_rows):
            r = i + 2
            ws_mc.cell(row=r, column=1, value=param)
            ws_mc.cell(row=r, column=2, value=val if isinstance(val, (int, float)) else str(val))

        _style_data(ws_mc, 2, len(mc_rows) + 1, len(headers_mc))
        _auto_width(ws_mc, len(headers_mc))

    # Сохраняем в bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
