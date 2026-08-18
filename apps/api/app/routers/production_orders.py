"""
Excel-импорт: трёхвкладочный формат (Заказы + BOM + Маршруты).
"""
import io
from datetime import date, datetime
from decimal import Decimal
from typing import List, Optional
from uuid import uuid4, UUID

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.deps import get_current_tenant_id
from app.models.production_order import ProductionOrder
from app.models.product_structure import ProductStructure
from app.models.routing import Routing, RoutingOperation
from app.models.nomenclature import Nomenclature
from app.models.resource import Resource
from app.schemas.production_order import (
    ProductionOrderCreate,
    ProductionOrderOut,
    ExcelImportResult,
    ImportValidationError,
)

router = APIRouter(prefix="/v1/production-orders", tags=["production-orders"])

# ── Helpers ────────────────────────────────────────────────────

def _parse_date(raw) -> Optional[date]:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(raw, default=Decimal("0")) -> Decimal:
    if raw is None or raw == "":
        return default
    try:
        return Decimal(str(raw).replace(",", "."))
    except Exception:
        return default


def _str(raw) -> str:
    if raw is None:
        return ""
    return str(raw).strip()


_NTYPE_MAP = {
    "assembly": "product",
    "semi_finished": "semi_finished",
    "material": "material",
    "phantom": "material",
    "service": "service",
}


def _normalize_name(name: str) -> str:
    """Нормализация имени для сопоставления: lower, trim, ё→е, коллапс пробелов."""
    s = (name or "").strip().lower().replace("ё", "е")
    return " ".join(s.split())


async def _get_or_create_nomenclature(
    db: AsyncSession,
    tenant_id,
    name: str,
    code: Optional[str],
    node_type: str,
    unit: str,
) -> tuple[Optional[UUID], str]:
    """Найти запись номенклатуры или создать новую.

    Приоритет сопоставления: code (ext_id) → нормализованное имя.
    Возвращает (nomenclature_id, action), где action ∈ {created, linked, skip}.
    """
    norm_name = _normalize_name(name)
    if not norm_name:
        return None, "skip"

    # 1. По внутреннему коду (ext_id узла как code)
    if code:
        row = (
            await db.execute(
                select(Nomenclature).where(
                    Nomenclature.tenant_id == tenant_id,
                    Nomenclature.code == code,
                )
            )
        ).scalars().first()
        if row:
            return row.id, "linked"

    # 2. По нормализованному имени (точное)
    row = (
        await db.execute(
            select(Nomenclature).where(
                Nomenclature.tenant_id == tenant_id,
                func.lower(Nomenclature.name) == norm_name,
            )
        )
    ).scalars().first()
    if row:
        return row.id, "linked"

    # 3. Создать новую
    n = Nomenclature(
        id=uuid4(),
        tenant_id=tenant_id,
        name=(name or "").strip(),
        code=code or None,
        ntype=_NTYPE_MAP.get(node_type, "material"),
        unit=unit or "pcs",
    )
    db.add(n)
    await db.flush()
    return n.id, "created"


# ── Карты русских значений (7-вкладочный формат) ──────────────

PRIORITY_MAP_RU = {
    "высокий": "high", "high": "high",
    "обычный": "normal", "normal": "normal",
    "низкий": "low", "low": "low",
    "критичный": "critical", "critical": "critical",
    "": "normal", None: "normal",
}

NODE_TYPE_MAP_RU = {
    "сборка": "assembly", "assembly": "assembly",
    "полуфабрикат": "semi_finished", "semi_finished": "semi_finished",
    "материал": "material", "material": "material",
    "фантом": "phantom", "phantom": "phantom",
}

RESOURCE_TYPE_MAP_RU = {
    "оборудование": "equipment", "equipment": "equipment",
    "персонал": "labor", "labor": "labor", "employee": "labor",
    "бригада": "team", "team": "team",
    "инструмент": "tool", "tool": "tool",
    "транспорт": "vehicle", "vehicle": "vehicle",
}


def _pick_sheet(wb, names):
    for n in names:
        if n in wb.sheetnames:
            return wb[n]
    return None


# ── POST /import ───────────────────────────────────────────────

@router.post("/import", response_model=ExcelImportResult)
async def import_excel(
    file: UploadFile = File(...),
    project_id: str = Form(None),
    tenant_id: str = Depends(get_current_tenant_id),
    db: AsyncSession = Depends(get_db),
):
    """
    Принимает .xlsx с тремя вкладками:
    1. Заказы (Orders)
    2. Состав (BOM)
    3. Маршруты (Routes)

    Создаёт ProductionOrder, BOM-узлы, техмаршруты и операции маршрутов.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Требуется файл .xlsx")

    content = await file.read()
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl не установлен на сервере")

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet_names = wb.sheetnames

    result = ExcelImportResult()

    # ── Вкладка 1: Заказы (Заказы / Orders / 2-Заказы) ─────
    ws = _pick_sheet(wb, ["Заказы", "Orders", "2-Заказы"])
    if ws is not None:
        result = await _import_orders(ws, project_id, tenant_id, db, result)

    # ── Вкладка 2: BOM (Состав / BOM / 3-BOM) ──────────────
    ws = _pick_sheet(wb, ["Состав", "BOM", "3-BOM"])
    if ws is not None:
        result = await _import_bom(ws, tenant_id, project_id, db, result)

    # ── Вкладка 4: Ресурсы (4-Ресурсы / Ресурсы) ────────
    ws = _pick_sheet(wb, ["4-Ресурсы", "Ресурсы"])
    if ws is not None:
        result = await _import_resources(ws, tenant_id, project_id, db, result)

    # ── Вкладка 3: Маршруты (5-Маршруты / Маршруты / Routes) ─
    # В 7-вкладочном формате у «5-Маршруты» сдвиг колонок +2 (есть Этап и Подразделение)
    ws = _pick_sheet(wb, ["5-Маршруты", "Маршруты", "Routes"])
    is_7tab = ws is not None and ws.title == "5-Маршруты"
    if ws is not None:
        result = await _import_routes(ws, tenant_id, db, result, project_id, is_7tab=is_7tab)

    # Предупреждение: нет ресурсов
    if result.resources_created == 0:
        result.warnings.append(
            "Ресурсы не указаны (вкладка «4-Ресурсы» пуста или отсутствует) — загрузка выполнена, но расчёт ресурсов и критической цепи будет ограничен. Ресурсы можно ввести вручную в справочнике."
        )

    # Проверка наличия маршрутов для make-узлов (сборка/полуфабрикат)
    if project_id:
        make_nodes = (await db.execute(
            select(ProductStructure).where(
                ProductStructure.tenant_id == tenant_id,
                ProductStructure.project_id == UUID(project_id),
                ProductStructure.node_type.in_(("assembly", "semi_finished")),
                ProductStructure.is_phantom == False,  # noqa: E712
                ProductStructure.routing_id.is_(None),
            )
        )).scalars().all()
        if make_nodes:
            result.warnings.append(
                f"У {len(make_nodes)} узлов сборки/полуфабриката нет маршрута — они не попадут в расчёт. Маршруты можно добавить вручную."
            )

    # Правила цепочки заказов: циклы parent_order_id и циклы BOM-связей
    await _validate_order_chain(project_id, tenant_id, db, result)

    await db.commit()
    return result


# ── Sheet parsers ──────────────────────────────────────────────


async def _import_orders(
    ws, project_id: Optional[str], tenant_id: str,
    db: AsyncSession, result: ExcelImportResult,
) -> ExcelImportResult:
    """Парсинг вкладки 'Заказы'.

    Колонки: ext_id | specification_name | specification_id | quantity |
             start_date | due_date | priority | client | parent_order_id(ext_id)
    """
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    ext_to_id: dict[str, UUID] = {}
    pending_parent: list[tuple[ProductionOrder, str, int]] = []  # (order, parent_ext_id, row)
    for i, row in enumerate(rows):
        if not row or not any(c for c in row):
            continue
        ext_id = _str(row[0]) if len(row) > 0 else ""
        spec_name = _str(row[1]) if len(row) > 1 else ""
        if not ext_id:
            result.errors.append(ImportValidationError(
                row=i + 2, sheet="Заказы", field="Код заказа",
                message="обязательное поле 'Код заказа' не заполнено",
            ))
            continue
        if not spec_name:
            result.errors.append(ImportValidationError(
                row=i + 2, sheet="Заказы", field="Продукт",
                message="обязательное поле 'Продукт' не заполнено",
            ))
            continue
        if not _str(row[3] if len(row) > 3 else None):
            result.errors.append(ImportValidationError(
                row=i + 2, sheet="Заказы", field="Кол-во",
                message="обязательное поле 'Кол-во' не заполнено",
            ))
            continue
        try:
            order = ProductionOrder(
                id=uuid4(),
                tenant_id=tenant_id,
                project_id=project_id or None,
                ext_id=ext_id,
                specification_name=spec_name,
                specification_id=_str(row[2]) if len(row) > 2 else None,
                quantity=_parse_decimal(row[3] if len(row) > 3 else 1, Decimal("1")),
                start_date=_parse_date(row[4] if len(row) > 4 else None),
                due_date=_parse_date(row[5] if len(row) > 5 else None),
                priority=PRIORITY_MAP_RU.get(_str(row[6]).lower(), "normal"),
                client=_str(row[7]) if len(row) > 7 else None,
                status="draft",
            )
            db.add(order)
            if order.ext_id:
                ext_to_id[order.ext_id] = order.id
            # parent_order_id — ext_id родителя, разрешаем после создания всех заказов
            parent_ext = _str(row[8]) if len(row) > 8 else ""
            if parent_ext:
                if parent_ext == ext_id:
                    result.errors.append(ImportValidationError(
                        row=i + 2, sheet="Заказы", field="Код заказа родителя",
                        message=f"заказ '{ext_id}' не может быть родителем самому себе",
                    ))
                else:
                    pending_parent.append((order, parent_ext, i + 2))
            result.orders_created += 1
        except Exception as e:
            result.errors.append(ImportValidationError(
                row=i + 2, sheet="Заказы", field="*",
                message=str(e),
            ))
    await db.flush()

    # Разрешаем parent_order_id: сначала импортированные заказы, затем уже существующие в БД проекта
    if project_id:
        db_orders = (await db.execute(
            select(ProductionOrder).where(
                ProductionOrder.tenant_id == tenant_id,
                ProductionOrder.project_id == UUID(project_id),
                ProductionOrder.ext_id.isnot(None),
            )
        )).scalars().all()
        for o in db_orders:
            if o.ext_id and o.ext_id not in ext_to_id:
                ext_to_id[o.ext_id] = o.id

    for order, parent_ext, row_no in pending_parent:
        pid = ext_to_id.get(parent_ext)
        if pid:
            order.parent_order_id = pid
        else:
            result.errors.append(ImportValidationError(
                row=row_no, sheet="Заказы", field="Код заказа родителя",
                message=f"код заказа родителя '{parent_ext}' не найден",
            ))
    await db.flush()
    return result


async def _import_bom(
    ws, tenant_id: str, project_id: Optional[str], db: AsyncSession, result: ExcelImportResult,
) -> ExcelImportResult:
    """Парсинг вкладки 'Состав (BOM)'.

    Колонки: spec_name | node_ext_id | parent_ext_id | node_type | nomenclature_name |
             unit | qty_per_parent | procurement_days | order_id(ext_id заказа-производителя)
    """
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # First pass: create all nodes
    node_map = {}  # ext_id → UUID

    # Карта заказов проекта: ext_id → order UUID (для привязки order_id)
    ext_to_order: dict[str, UUID] = {}
    orders_q = (await db.execute(
        select(ProductionOrder).where(
            ProductionOrder.tenant_id == tenant_id,
            ProductionOrder.project_id == (UUID(project_id) if project_id else None),
        )
    )).scalars().all()
    for o in orders_q:
        if o.ext_id:
            ext_to_order[o.ext_id] = o.id

    pending_order_link: list[tuple[ProductStructure, str, int]] = []  # (node, order_ext_id, row)

    for i, row in enumerate(rows):
        if not row or not any(c for c in row):
            continue
        spec_name = _str(row[0]) if len(row) > 0 else ""
        node_ext_id = _str(row[1]) if len(row) > 1 else ""
        nomenclature_name = _str(row[4]) if len(row) > 4 else ""
        if not spec_name:
            result.errors.append(ImportValidationError(
                row=i + 2, sheet="BOM", field="Спецификация",
                message="обязательное поле 'Спецификация' не заполнено",
            ))
            continue
        if not node_ext_id:
            result.errors.append(ImportValidationError(
                row=i + 2, sheet="BOM", field="Узел ID",
                message="обязательное поле 'Узел ID' не заполнено",
            ))
            continue
        if not nomenclature_name:
            result.errors.append(ImportValidationError(
                row=i + 2, sheet="BOM", field="Номенклатура",
                message="обязательное поле 'Номенклатура' не заполнено",
            ))
            continue
        if not _str(row[3] if len(row) > 3 else None):
            result.errors.append(ImportValidationError(
                row=i + 2, sheet="BOM", field="Тип",
                message="обязательное поле 'Тип' не заполнено",
            ))
            continue
        if not _str(row[6] if len(row) > 6 else None):
            result.errors.append(ImportValidationError(
                row=i + 2, sheet="BOM", field="Норма на 1",
                message="обязательное поле 'Норма на 1' не заполнено",
            ))
            continue
        try:
            node_type_raw = _str(row[3]).lower() if len(row) > 3 else "material"
            node_type = NODE_TYPE_MAP_RU.get(node_type_raw, "material")
            is_phantom = (node_type_raw in ("phantom", "фантом"))
            unit_str = _str(row[5]) if len(row) > 5 else "pcs"

            nref_id, nact = await _get_or_create_nomenclature(
                db, tenant_id, nomenclature_name, node_ext_id, node_type, unit_str
            )
            if nact == "created":
                result.nomenclature_created += 1
            elif nact == "linked":
                result.nomenclature_linked += 1

            node = ProductStructure(
                id=uuid4(),
                tenant_id=tenant_id,
                project_id=UUID(project_id) if project_id else None,
                nomenclature_id=node_ext_id,
                nomenclature_ref_id=nref_id,
                nomenclature_name=nomenclature_name,
                node_type=node_type,
                is_phantom=is_phantom,
                quantity_per_parent=_parse_decimal(row[6] if len(row) > 6 else 1, Decimal("1")),
                unit=unit_str,
                procurement_lead_time_days=_parse_decimal(row[7] if len(row) > 7 else None, None),
                is_make_or_buy="make" if node_type in ("assembly","semi_finished") else "buy",
            )
            # Используем id родительской спецификации (spec) как путь
            node.path = f"{spec_name}/{node_ext_id}" if spec_name else node_ext_id
            db.add(node)
            node_map[node_ext_id] = node.id

            # order_id — ext_id заказа-производителя (куст заказов)
            order_ext = _str(row[8]) if len(row) > 8 else ""
            if order_ext:
                pending_order_link.append((node, order_ext, i + 2))

            result.bom_nodes_created += 1
        except Exception as e:
            result.errors.append(ImportValidationError(
                row=i + 2, sheet="BOM", field="*",
                message=str(e),
            ))

    await db.flush()

    # Привязываем order_id (ext_id → order UUID)
    for node, order_ext, row_no in pending_order_link:
        oid = ext_to_order.get(order_ext)
        if oid:
            node.order_id = oid
        else:
            result.errors.append(ImportValidationError(
                row=row_no, sheet="BOM", field="Код заказа",
                message=f"код заказа '{order_ext}' не найден",
            ))

    # Second pass: set parent relationships
    for i, row in enumerate(rows):
        if not row or not any(c for c in row):
            continue
        try:
            node_ext_id = _str(row[1]) if len(row) > 1 else ""
            parent_ext_id = _str(row[2]) if len(row) > 2 else ""
            if parent_ext_id and parent_ext_id in node_map and node_ext_id in node_map:
                node = await db.get(ProductStructure, node_map[node_ext_id])
                if node:
                    node.parent_id = node_map[parent_ext_id]
        except Exception:
            pass

    await db.flush()
    return result


async def _import_resources(
    ws, tenant_id: str, project_id: Optional[str], db: AsyncSession, result: ExcelImportResult,
) -> ExcelImportResult:
    """Парсинг вкладки 'Ресурсы'.

    Колонки: ID ресурса | Название | Тип | Подразделение | Доступно | Ед.
    """
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for i, row in enumerate(rows):
        if not row or not any(c for c in row):
            continue
        name = _str(row[1]) if len(row) > 1 else ""
        if not name:
            result.errors.append(ImportValidationError(
                row=i + 2, sheet="Ресурсы", field="Название",
                message="обязательное поле 'Название' не заполнено",
            ))
            continue
        rtype_raw = _str(row[2]).lower() if len(row) > 2 else "equipment"
        rtype = RESOURCE_TYPE_MAP_RU.get(rtype_raw, "equipment")
        available = _parse_decimal(row[4] if len(row) > 4 else 1, Decimal("1"))
        unit = _str(row[5]) if len(row) > 5 else "pcs"
        try:
            res = Resource(
                id=uuid4(),
                tenant_id=tenant_id,
                project_id=UUID(project_id) if project_id else None,
                name=name,
                resource_type=rtype,
                capacity_per_unit=available,
                capacity_unit="hour",
                unit=unit,
                ext_id=_str(row[0]) if len(row) > 0 else None,
            )
            db.add(res)
            result.resources_created += 1
        except Exception as e:
            result.errors.append(ImportValidationError(
                row=i + 2, sheet="Ресурсы", field="*",
                message=str(e),
            ))
    await db.flush()
    return result


def _find_cycles_multi(graph: dict[UUID, set[UUID]]) -> list[list[UUID]]:
    """Поиск уникальных циклов в ориентированном мультиграфе (DFS)."""
    cycles: list[list[UUID]] = []
    seen: set[frozenset] = set()
    for start in graph:
        stack: list[tuple[UUID, list[UUID], set[UUID]]] = [(start, [start], {start})]
        while stack:
            node, path, path_set = stack.pop()
            for nxt in graph.get(node, set()):
                if nxt in path_set:
                    idx = path.index(nxt)
                    cyc = path[idx:] + [nxt]
                    key = frozenset(cyc)
                    if key not in seen:
                        seen.add(key)
                        cycles.append(cyc)
                elif nxt in graph:
                    stack.append((nxt, path + [nxt], path_set | {nxt}))
    return cycles


def _build_bom_link_graph(
    orders: list[ProductionOrder], nodes: list[ProductStructure],
) -> tuple[dict[UUID, set[UUID]], dict[UUID, str]]:
    """Граф связей BOM: владелец узла (по спецификации из path) → заказ-производитель узла.

    Возвращает (graph, ext_by_id).
    """
    order_by_spec: dict[str, ProductionOrder] = {}
    for o in orders:
        # «Спецификация» в BOM = specification_id заказа (колонка «Спецификация»), возможен и name
        if o.specification_id:
            order_by_spec[o.specification_id.strip()] = o
        if o.specification_name:
            order_by_spec.setdefault(o.specification_name.strip(), o)
    graph: dict[UUID, set[UUID]] = {}
    ext_by_id: dict[UUID, str] = {}
    for o in orders:
        graph.setdefault(o.id, set())
        ext_by_id[o.id] = o.ext_id or str(o.id)[:8]
    for n in nodes:
        if not n.order_id or n.order_id not in graph:
            continue
        spec = ""
        if n.path and "/" in n.path:
            spec = n.path.rsplit("/", 1)[0]
        owner = order_by_spec.get(spec.strip())
        if owner:
            if n.order_id == owner.id:
                continue  # ссылка на свой же заказ — допустима, цикла не создаёт
            graph[owner.id].add(n.order_id)
    return graph, ext_by_id


async def _validate_order_chain(
    project_id: Optional[str], tenant_id: str, db: AsyncSession, result: ExcelImportResult,
) -> None:
    """Проверка цепочки заказов после импорта.

    Правила:
      - в цепочке parent_order_id не должно быть циклов (нельзя ссылаться на себя или на потомка);
      - в связях BOM (order_id на узлах) не должно быть циклов между заказами.
    """
    if not project_id:
        return
    orders = (await db.execute(
        select(ProductionOrder).where(
            ProductionOrder.tenant_id == tenant_id,
            ProductionOrder.project_id == UUID(project_id),
        )
    )).scalars().all()
    by_id = {o.id: o for o in orders}
    ext_by_id = {o.id: o.ext_id or str(o.id)[:8] for o in orders}

    # 1) Циклы в цепочке parent_order_id
    parent_graph: dict[UUID, set[UUID]] = {o.id: set() for o in orders}
    for o in orders:
        if o.parent_order_id and o.parent_order_id in by_id:
            parent_graph[o.id].add(o.parent_order_id)
    for cyc in _find_cycles_multi(parent_graph):
        chain = " → ".join(ext_by_id.get(c, str(c)[:8]) for c in cyc)
        result.errors.append(ImportValidationError(
            row=0, sheet="Заказы", field="Код заказа родителя",
            message=f"обнаружен цикл в цепочке заказов: {chain}",
        ))

    # 2) Циклы в связях BOM (order_id на узлах)
    nodes = (await db.execute(
        select(ProductStructure).where(
            ProductStructure.tenant_id == tenant_id,
            ProductStructure.project_id == UUID(project_id),
            ProductStructure.order_id.isnot(None),
        )
    )).scalars().all()
    bom_graph, _ = _build_bom_link_graph(orders, nodes)
    for cyc in _find_cycles_multi(bom_graph):
        chain = " → ".join(ext_by_id.get(c, str(c)[:8]) for c in cyc)
        result.errors.append(ImportValidationError(
            row=0, sheet="BOM", field="Код заказа",
            message=f"обнаружен цикл в связях BOM по цепочке заказов: {chain}",
        ))


async def _import_routes(
    ws, tenant_id: str, db: AsyncSession, result: ExcelImportResult,
    project_id: Optional[str] = None, is_7tab: bool = False,
) -> ExcelImportResult:
    """Парсинг вкладки 'Маршруты'.

    В 7-вкладочном формате колонки сдвинуты на +2 (есть «Этап» и «Подразделение»):
      Длит.ч → col 6, Предш.оп. → col 7, Доп.материал → col 8, Расход → col 9, Вых.год. → col 10.
    """
    if is_7tab:
        # 7-вкладочный: длительность на 6, предш.оп. на 7, материалы 8-9, вых.год. 10
        c_dur, c_out, c_pred, c_mat, c_qty, c_yield = 6, None, 7, 8, 9, 10
    else:
        c_dur, c_out, c_pred, c_mat, c_qty, c_yield = 4, 5, 6, 7, 8, 9
    # Группируем строки по node_id
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    routings_by_node = {}  # node_ext_id → [(row_idx, row_data)]
    ops_without_resource = 0

    for i, row in enumerate(rows):
        if not row or not any(c for c in row):
            continue
        node_id = _str(row[0]) if len(row) > 0 else ""
        if node_id:
            routings_by_node.setdefault(node_id, []).append((i, row))

    # Находим BOM-узлы по ext_id
    for node_ext_id, node_rows in routings_by_node.items():
        try:
            stmt = select(ProductStructure).where(
                ProductStructure.nomenclature_id == node_ext_id,
                ProductStructure.tenant_id == tenant_id,
                ProductStructure.project_id == (UUID(project_id) if project_id else None),
            )
            res = await db.execute(stmt)
            bom_node = res.scalar_one_or_none()
            if not bom_node:
                continue

            # Создаём Routing
            routing = Routing(
                id=uuid4(),
                tenant_id=tenant_id,
                name=f"Маршрут: {bom_node.nomenclature_name}",
                product_node_id=bom_node.id,
                variant="Основной",
                is_default=True,
            )
            db.add(routing)
            await db.flush()
            result.routings_created += 1

            # Создаём RoutingOperation для каждой строки
            for row_idx, row in node_rows:
                seq_val = row[1] if len(row) > 1 else None
                name_val = _str(row[2]) if len(row) > 2 else ""
                dur_val = row[c_dur] if len(row) > c_dur else None
                if not seq_val or not name_val or dur_val is None or _str(dur_val) == "":
                    result.errors.append(ImportValidationError(
                        row=row_idx + 2, sheet="Маршруты", field="№ оп./Операция/Длит.",
                        message="обязательные поля '№ оп.', 'Операция', 'Длит.,ч' не заполнены",
                    ))
                    continue
                res_name = _str(row[3]) if len(row) > 3 else ""
                if not res_name:
                    ops_without_resource += 1
                try:
                    rop = RoutingOperation(
                        id=uuid4(),
                        routing_id=routing.id,
                        sequence_number=int(seq_val) if seq_val else 1,
                        name=name_val,
                        duration_hours=_parse_decimal(dur_val),
                        setup_hours=Decimal("0"),
                        resource_type_id=res_name or None,
                        output_product=(
                            _str(row[c_out]) if c_out is not None and len(row) > c_out and _str(row[c_out]) else None
                        ),
                        output_quantity=_parse_decimal(Decimal("1")),
                        yield_rate=_parse_decimal(row[c_yield] if len(row) > c_yield else 1, Decimal("1")),
                        predecessors=_str(row[c_pred]) if len(row) > c_pred and row[c_pred] else None,
                    )
                    db.add(rop)

                    # Дополнительные материалы
                    extra_mat = _str(row[c_mat]) if len(row) > c_mat else ""
                    extra_qty = float(row[c_qty]) if len(row) > c_qty and row[c_qty] else 0
                    if extra_mat and extra_qty > 0:
                        import json
                        rop.input_materials = json.dumps([{
                            "name": extra_mat,
                            "qty": extra_qty,
                        }])

                    result.routing_ops_created += 1
                except Exception as e:
                    result.errors.append(ImportValidationError(
                        row=row_idx + 2, sheet="Маршруты", field="*",
                        message=str(e),
                    ))

            # Линкуем routing_id к BOM-узлу
            bom_node.routing_id = routing.id

        except Exception as e:
            result.errors.append(ImportValidationError(
                row=0, sheet="Маршруты", field="node_id",
                message=f"Узел {node_ext_id}: {e}",
            ))

    if ops_without_resource > 0:
        result.warnings.append(
            f"В {ops_without_resource} операциях не указан ресурс — загрузка выполнена, но расчёт ресурсов будет неполным. Ресурсы можно добавить вручную в справочнике."
        )

    await db.flush()
    return result


# ── GET / ──────────────────────────────────────────────────────

@router.get("/", response_model=list[ProductionOrderOut])
async def list_orders(
    project_id: Optional[str] = None,
    tenant_id: str = Depends(get_current_tenant_id),
    db: AsyncSession = Depends(get_db),
):
    """Список заказов на производство, опционально отфильтрованный по проекту."""
    stmt = select(ProductionOrder).where(ProductionOrder.tenant_id == tenant_id)
    if project_id:
        stmt = stmt.where(ProductionOrder.project_id == UUID(project_id))
    stmt = stmt.order_by(ProductionOrder.created_at.desc())
    res = await db.execute(stmt)
    orders = res.scalars().all()
    return [_order_to_out(o) for o in orders]


async def _resolve_parent_order(
    value: Optional[str],
    project_id: Optional[str],
    tenant_id: str,
    db: AsyncSession,
) -> Optional[UUID]:
    """Разрешает parent_order_id: принимает UUID или ext_id заказа.

    Возвращает UUID родительского заказа или None.
    """
    if not value:
        return None
    v = value.strip()
    # 1. Это валидный UUID?
    try:
        pid = UUID(v)
        # Проверим, что такой заказ существует
        exists = (await db.execute(
            select(ProductionOrder.id).where(ProductionOrder.id == pid)
        )).scalar_one_or_none()
        return pid if exists else None
    except ValueError:
        pass
    # 2. Это ext_id — ищем заказ по ext_id в рамках проекта/арендатора
    stmt = select(ProductionOrder).where(
        ProductionOrder.tenant_id == tenant_id,
        ProductionOrder.ext_id == v,
    )
    if project_id:
        stmt = stmt.where(ProductionOrder.project_id == UUID(project_id))
    res = await db.execute(stmt)
    order = res.scalars().first()
    return order.id if order else None


# ── POST / ─────────────────────────────────────────────────────

@router.post("/", response_model=ProductionOrderOut, status_code=201)
async def create_order(
    project_id: str,
    body: ProductionOrderCreate,
    tenant_id: str = Depends(get_current_tenant_id),
    db: AsyncSession = Depends(get_db),
):
    """Создать один заказ на производство."""
    from uuid import UUID
    parent_id = await _resolve_parent_order(body.parent_order_id, project_id, tenant_id, db)
    order = ProductionOrder(
        id=uuid4(),
        tenant_id=tenant_id,
        project_id=UUID(project_id),
        ext_id=body.ext_id,
        specification_name=body.specification_name,
        quantity=body.quantity,
        unit=body.unit,
        start_date=body.start_date,
        due_date=body.due_date,
        priority=body.priority,
        client=body.client,
        notes=body.notes,
        parent_order_id=parent_id,
    )
    db.add(order)
    await db.commit()
    await db.refresh(order)
    return _order_to_out(order)


# ── GET /check-duplicate ───────────────────────────────────────
# (должен быть объявлен ДО /{order_id}, иначе "check-duplicate" попадёт в UUID-парсер)

@router.get("/check-duplicate")
async def check_duplicate(
    project_id: str,
    ext_id: Optional[str] = None,
    specification_name: Optional[str] = None,
    tenant_id: str = Depends(get_current_tenant_id),
    db: AsyncSession = Depends(get_db),
):
    """Проверить возможный дубликат заказа по ext_id + specification_name.

    Возвращает список существующих заказов с совпадающими полями (в рамках проекта).
    """
    if not ext_id and not specification_name:
        return {"duplicate": False, "existing": []}
    stmt = select(ProductionOrder).where(
        ProductionOrder.tenant_id == tenant_id,
        ProductionOrder.project_id == UUID(project_id),
    )
    if specification_name:
        stmt = stmt.where(ProductionOrder.specification_name == specification_name)
    if ext_id:
        stmt = stmt.where(ProductionOrder.ext_id == ext_id)
    res = await db.execute(stmt)
    dups = res.scalars().all()
    return {
        "duplicate": len(dups) > 0,
        "existing": [
            {
                "id": str(o.id),
                "ext_id": o.ext_id,
                "specification_name": o.specification_name,
                "status": o.status,
            }
            for o in dups
        ],
    }


# ── GET /{id} ──────────────────────────────────────────────────

@router.get("/{order_id}", response_model=ProductionOrderOut)
async def get_order(
    order_id: str,
    tenant_id: str = Depends(get_current_tenant_id),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(ProductionOrder).where(
        ProductionOrder.id == UUID(order_id),
        ProductionOrder.tenant_id == tenant_id,
    )
    res = await db.execute(stmt)
    order = res.scalar_one_or_none()
    if not order:
        raise HTTPException(404, "Заказ не найден")
    return _order_to_out(order)


def _order_to_out(o: ProductionOrder) -> ProductionOrderOut:
    return ProductionOrderOut(
        id=str(o.id),
        tenant_id=str(o.tenant_id),
        project_id=str(o.project_id) if o.project_id else None,
        ext_id=o.ext_id,
        specification_id=o.specification_id,
        specification_name=o.specification_name,
        quantity=o.quantity,
        unit=o.unit,
        start_date=o.start_date,
        due_date=o.due_date,
        priority=o.priority,
        client=o.client,
        notes=o.notes,
        status=o.status,
        group_id=str(o.group_id) if o.group_id else None,
        pool_id=str(o.pool_id) if o.pool_id else None,
        parent_order_id=str(o.parent_order_id) if o.parent_order_id else None,
        exploded_at=o.exploded_at,
        operations_created=o.operations_created,
        created_at=o.created_at,
    )


# ── POST /{id}/expand ──────────────────────────────────────────

from app.services.bom_explosion import ExplodedOperation, ExplodedDependency
from app.models.operation import Operation, OperationDependency


@router.post("/{order_id}/expand")
async def expand_order(
    order_id: str,
    tenant_id: str = Depends(get_current_tenant_id),
    db: AsyncSession = Depends(get_db),
):
    """Разворачивает BOM-спецификацию заказа в CPM-операции."""
    stmt = select(ProductionOrder).where(
        ProductionOrder.id == UUID(order_id),
        ProductionOrder.tenant_id == tenant_id,
    )
    res = await db.execute(stmt)
    order = res.scalar_one_or_none()
    if not order:
        raise HTTPException(404, "Заказ не найден")

    if not order.project_id:
        raise HTTPException(400, "У заказа не указан проект (project_id). Сначала импортируйте Excel и укажите project_id.")

    # Запускаем развёртку через реальный движок
    from app.routers.bom import run_explosion
    from app.schemas.bom import BOMExplodeAndSaveRequest

    body = BOMExplodeAndSaveRequest(project_quantity=order.quantity)
    result = await run_explosion(
        db=db,
        tenant_id=tenant_id,
        project_id=order.project_id,
        body=body,
    )

    if result.warnings:
        import logging
        for w in result.warnings:
            logging.warning(f"BOM expand warning: {w}")

    # Сохраняем операции
    op_map = {}  # temp_id → Operation.id
    for eop in result.operations:
        op = Operation(
            id=uuid4(),
            tenant_id=tenant_id,
            project_id=order.project_id,
            name=eop.name,
            duration_base=eop.duration_base,
            duration_unit=eop.duration_unit,
            setup_time=eop.setup_time,
            teardown_time=eop.teardown_time,
            operation_type=eop.operation_type,
            output_product=eop.output_product,
            output_quantity=eop.output_quantity,
            yield_rate=eop.yield_rate,
            input_materials=eop.input_materials,
            supplier_id=eop.supplier_id,
            is_milestone=eop.is_milestone,
        )
        db.add(op)
        await db.flush()
        op_map[eop.temp_id] = op.id

    # Сохраняем зависимости
    for dep in result.dependencies:
        if dep.predecessor_temp_id in op_map and dep.successor_temp_id in op_map:
            d = OperationDependency(
                id=uuid4(),
                predecessor_id=op_map[dep.predecessor_temp_id],
                successor_id=op_map[dep.successor_temp_id],
                dependency_type=dep.dependency_type,
                lag_time=dep.lag_hours,
                lag_unit="hour",
            )
            db.add(d)

    # Обновляем заказ
    from datetime import datetime, timezone
    order.status = "planned"
    order.exploded_at = datetime.now(timezone.utc)
    order.operations_created = len(result.operations)

    await db.commit()

    return {
        "order_id": str(order.id),
        "status": order.status,
        "operations_created": len(result.operations),
        "dependencies_created": len(result.dependencies),
        "materials_required": len(result.materials),
        "warnings": result.warnings,
    }


# ── PUT /{id} ─────────────────────────────────────────────────

@router.put("/{order_id}", response_model=ProductionOrderOut)
async def update_order(
    order_id: str,
    body: ProductionOrderCreate,
    tenant_id: str = Depends(get_current_tenant_id),
    db: AsyncSession = Depends(get_db),
):
    """Обновить поля заказа."""
    stmt = select(ProductionOrder).where(
        ProductionOrder.id == UUID(order_id),
        ProductionOrder.tenant_id == tenant_id,
    )
    res = await db.execute(stmt)
    order = res.scalar_one_or_none()
    if not order:
        raise HTTPException(404, "Заказ не найден")
    for field in ("specification_name", "ext_id", "unit", "priority", "client", "notes", "status"):
        v = getattr(body, field, None)
        if v is not None:
            setattr(order, field, v)
    if body.quantity is not None:
        order.quantity = body.quantity
    if body.start_date is not None:
        order.start_date = body.start_date
    if body.due_date is not None:
        order.due_date = body.due_date
    if body.parent_order_id is not None:
        order.parent_order_id = await _resolve_parent_order(
            body.parent_order_id, str(order.project_id) if order.project_id else None,
            tenant_id, db
        )
    await db.commit()
    await db.refresh(order)
    return _order_to_out(order)


# ── PATCH /{id}/move ───────────────────────────────────────────

from pydantic import BaseModel as PydanticBase


class OrderMoveRequest(PydanticBase):
    group_id: Optional[str] = None
    pool_id: Optional[str] = None


@router.patch("/{order_id}/move", response_model=ProductionOrderOut)
async def move_order(
    order_id: str,
    body: OrderMoveRequest,
    tenant_id: str = Depends(get_current_tenant_id),
    db: AsyncSession = Depends(get_db),
):
    """Переместить заказ в группу/пул или убрать из них.
    
    Передайте group_id или pool_id (не оба сразу).
    Передайте оба null чтобы убрать заказ из группы/пула и вернуть в корень.
    """
    stmt = select(ProductionOrder).where(
        ProductionOrder.id == UUID(order_id),
        ProductionOrder.tenant_id == tenant_id,
    )
    res = await db.execute(stmt)
    order = res.scalar_one_or_none()
    if not order:
        raise HTTPException(404, "Заказ не найден")
    order.group_id = UUID(body.group_id) if body.group_id else None
    order.pool_id = UUID(body.pool_id) if body.pool_id else None
    await db.commit()
    await db.refresh(order)
    return _order_to_out(order)


# ── DELETE /{id} ──────────────────────────────────────────────

@router.delete("/{order_id}", status_code=204)
async def delete_order(
    order_id: str,
    tenant_id: str = Depends(get_current_tenant_id),
    db: AsyncSession = Depends(get_db),
):
    """Удалить заказ на производство."""
    stmt = select(ProductionOrder).where(
        ProductionOrder.id == UUID(order_id),
        ProductionOrder.tenant_id == tenant_id,
    )
    res = await db.execute(stmt)
    order = res.scalar_one_or_none()
    if not order:
        raise HTTPException(404, "Заказ не найден")
    await db.delete(order)
    await db.commit()
